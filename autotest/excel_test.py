#通过excel，编写测试用例，将数据传入到代码中去
import openpyxl
import os
#仅能读取xlsx
from openpyxl import load_workbook
wb = load_workbook("testdata.xlsx")

#定位表单，表单名为"test1"
#sheet= wb.get_sheet_by_name("test1") #deprecated，过时的方法
#使用新方法
sheet = wb["testdata"]

#定位单元格并取值
#3行2列的值
res = sheet.cell(3,2).value
#修改单元格的值
#若写excel时，需要保证excel未被打开
# sheet.cell(1,1).value = "呵呵(A1)"
#修改需要保存，否则修改无效
# wb.save("testdata.xlsx")

test_data = []
#遍历excel中的所有数据
for i in range(2,7):#行
    datas=[]
    for j in range(1,6):#列
        datas.append(sheet.cell(i,j).value)
    test_data.append(datas)
print(test_data)

if __name__ =='__main__':
   # path0 = os.path.abspath(__file__) #返回绝对路径
   # path1 = os.path.realpath(__file__)  #返回相对路径
    #path2 = os.path.dirname(__file__)#返回文件路径
    path3 = os.path.realpath(os.path.dirname(__file__))
    excel_path = "testdata.xlsx"
    path5 = os.path.join(path3,excel_path)

   #准确性较高
    path4 = os.path.abspath(os.path.dirname(__file__))
    #rootPath0 = os.path.split(path4)
    rootPath1 = os.path.split(path4)[0] #获取根目录

    #print("path0",path0)
    #print("path1",path1)
   # print("path2",path2)
    print("path3",path3)
    print("path4",path4)
    print("path5",path5)
    #print("rootPath0",rootPath0)
    print("rootPath1",rootPath1)