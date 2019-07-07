#读取测试用例的数据
from openpyxl import load_workbook
from API_Frame_Work.common.get_path import Get_Path
from API_Frame_Work.common.read_config import read_config
import os

class manage_excel():

    def __init__(self):
        self.testcase_file = read_config().config_Test_Data()["file_name"] #获取文件名
        self.testcase_sheet = read_config().config_Test_Data()["sheet_name"]#获取表名
        self.testcasepath = os.path.join(Get_Path().get_test_data_path(),self.testcase_file)#获取表的路径

        self.testresult_file = read_config().config_Test_Result()["file_name"]
        self.testresult_sheet = read_config().config_Test_Result()["sheet_name"]
        self.testresultpath =os.path.join(Get_Path().get_test_result_path(),self.testresult_file)
        print(self.testresultpath)

    def read_excel(self):
        #打开文件
        wb = load_workbook(self.testcasepath)
        #定位到对应的sheet
        ws = wb[self.testcase_sheet]
        #获取最大行数，列数
        max_row = ws.max_row
        max_column = ws.max_column
        #输出数据
        test_data = []
        for row in range(2,max_row): #测试数据的取值行数从2开始
            sub_data = {}
            for column in range(1,max_column-1):#测试数据的取值列数需要减1行
                        #遍历表头的title
                sub_data[ws.cell(1,column).value] = ws.cell(row,column).value
            test_data.append(sub_data)
        return test_data
    #将结果写回到excel中去
    def write_back(self,row,ActualResult,TestResult):
        wb = load_workbook(self.testresultpath)
        ws = wb[self.testresult_sheet]
        ws.cell(row,ws.max_column-1).value = ActualResult
        ws.cell(row,ws.max_column).value = TestResult
        wb.save(self.testresultpath)


if __name__ =='__main__':
    manage_excel = manage_excel()
    print(manage_excel.read_excel())
    print(manage_excel.write_back(2,101010,'Failed'))










