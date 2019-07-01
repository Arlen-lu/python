import os
from openpyxl import  Workbook
from openpyxl import load_workbook

class GetPath():

    def __init__(self,test_excel):
        self.path = os.path.realpath(os.path.dirname(__file__))#获取当前文件的上一级目录
        self.rootPaht = os.path.split(self.path)[0] #获取当前文件的艮目录
        #实例化excel文件的名称
        self.test_excel = test_excel


    def AddExcel_path(self):
        excel_path = os.path.join(self.path,self.test_excel)
        return(excel_path)

class ManageExcel():

    def __init__(self,excel_path):
        self.excel_path = excel_path

    def New_excel(self):
        wb = Workbook() #新建excel
        ws2 = wb.active #获取第一个sheet，无()
        ws2.title = "Mysheet2" #将默认存在的sheet改名为Mysheet0
        ws0 = wb.create_sheet("Mysheet0",0) #新建表create_sheet(title=[表名]，index=[位置])
        ws1 = wb.create_sheet(index=1) #在1位置新建sheet
        ws1.title = "Mysheet1"#修改表名为Mysheet1
        wb.save(self.excel_path) #保存操作，即新建excel

    def Data_In(self):
        #打开文件
        wb = load_workbook(self.excel_path)
        wb_names = wb.sheetnames  #获取所有的sheet names，并以list方式展示
        # ws = wb.active  #默认定位到首个sheet
        ws0 = wb[wb_names[0]]  #定位excel中的首个sheet
        #在第一个sheet中写入数据
        ws0['A1'] = 1
        #ws0.append([0,1,2,3,4,5]) #添加行数据
        ws0['B2'] = 2
        ws0['C3'] = 3
        ws0['D4'] = 4
        ws0['E5'] = 5
        ws0.cell(5,5,10) #ws0.cell(行，列，值)

        wb.save(self.excel_path)

    def Check_data(self):
        #打开文件
        wb = load_workbook(self.excel_path)
        wb_names = wb.sheetnames
        ws0 = wb[wb_names[0]]
        #批量操作,单行单列都为元组
        #操作单列
        # for column in ws0['A']: #定位到A列
        #     return(column.value)
        #     # ws0['A'].append([lambda a:a in range(1,10)]) #单列为元组，无法利用append添加数据
        #     # ws0[cell] = 1
        # #操作单行
        # for row in ws0[4]:  #定位到第4行
        #     return(row.value)
        # #所有行，列
        # for row in ws0.rows:
        #     return(row)
        # for column in ws0.columns:
        #     return(column)

        #查看1行1列到3行3列所有的数据
        for row in ws0.iter_rows(min_row=1,min_col=1,max_row=3,max_col=3):
            #print("row",row,type(row)) #row为tuple类型
            for cell in row:
                return(cell.value)





if __name__ =='__main__':
    test_excel="test.xlsx"
    getpath = GetPath(test_excel)
    path = getpath.AddExcel_path()
    manage = ManageExcel(path)
    # manage.New_excel()
    manage.Data_In()
    manage.Check_data()