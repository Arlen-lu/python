#获取excel文件路径，获取TestCase的数据，并返回测试结果
import openpyxl
from openpyxl import load_workbook
import os
from read_config import ReadConfig

class GetExcel():

    def __init__(self,test_excel):
        self.path = os.path.realpath(os.path.dirname('__file__'))
        self.rootPath = os.path.split(self.path)[0]
        self.test_excel = test_excel

    def Get_Excel_Path(self):
        excel_path = os.path.join(self.path,self.test_excel)
        return excel_path

class DoExcel():

    def __init__(self,excel_path,excel_title):
        self.wb = load_workbook(excel_path)
        self.sheet = self.wb[excel_title]
        self.excel_path = excel_path


#版本一，列表嵌套列表
    def do_excel1(self):
        test_data = []
        for i in range(2,7):#行
            sub_data=[]
            for j in range(1,6):#列
                sub_data.append(self.sheet.cell(i,j).value)
            test_data.append(sub_data)
        return(test_data)

#版本二，列表嵌套字典
    def do_excel2(self,button,case_id_list):
        test_data = []
        #先获取表头
        header = []
        for i in range(1,6):
            header.append(self.sheet.cell(1,i).value)
        # print("Header{0}".format(header))
        for i in range(2,7):#行
            sub_data = {}
            for j in range(1,6):#列
                    #表头对应数据
                    sub_data[header[j-1]] = self.sheet.cell(i,j).value
            test_data.append(sub_data)


        #将配置文件放在这
        final_data = []#存储最终要执行的数据
        if button =='on':#跑所有的数据
            final_data = test_data
        else:#否则就执行case_id_list中的case号
            for item in test_data:
                if item["id"] in case_id_list:
                    final_data.append(item)
        return(final_data)

    #将结果写回到excel中
    def Write_Back(self,row,ActualResult,TestResult):
        self.sheet.cell(row,6).value = ActualResult
        self.sheet.cell(row,7).value = TestResult
        self.wb.save(self.excel_path)




if __name__ =='__main__':
    #获取配置文件中的值
    button = ReadConfig().read_config("case.config",'FLAG','button')
    case_id_list = eval(ReadConfig().read_config("case.config",'FLAG','cse_id_list'))
    test_excel = "testdata.xlsx"
    excel_title = "testdata"
    excel_path = GetExcel(test_excel).Get_Excel_Path()
    print(type(excel_path))

    test_data = DoExcel(excel_path,excel_title).do_excel1()
    print("获取到的数据{0}".format(test_data))
    test_data = DoExcel(excel_path,excel_title).do_excel2(button,case_id_list)
    print("获取到的数据{0}".format(test_data))
